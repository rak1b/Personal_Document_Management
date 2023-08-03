from coreapp.helper import *
from coreapp.models import *
import uuid
from datetime import datetime
from django.core.files.storage import FileSystemStorage
from openpyxl_image_loader import SheetImageLoader

from django.core.files.base import ContentFile
from django.conf import settings
from openpyxl import load_workbook
import json
import pathlib

class BulkImportUtils:
    def set_file_path(self, filepath):
        self.csv_filepath = "media/" + filepath + "/csv/"
        self.img_filepath = filepath + "/images/"

    def create_document(self, image):
        document = Document.objects.create(
            document=image, owner=self.request.user, doc_type=0
        )
        return document

    def bulk_import(self, uploaded_file_path):
        pass

    def get_file_name(self, prefix="Product"):
        return (
            prefix
            + str(uuid.uuid4())
            + datetime.now().strftime("%Y_%m_%d-%I_%M_%S_%p")
            + ".xlsx"
        )

    def upload_file(self, file, file_path, prefix="Product"):
        self.set_file_path(file_path)
        fs = FileSystemStorage(location=self.csv_filepath)
        content = file.read()
        file_content = ContentFile(content)
        file_name = fs.save(self.get_file_name(prefix=prefix), file_content)
        uploaded_file_path = fs.path(file_name)
        return uploaded_file_path


class ProductBulkImport(BulkImportUtils):
    images = []
    result = []
    heading = []

    def get_or_create_sub_main_category(self, sub_category, main_category):
        from inventory.models import Category
        sub_category_obj, is_created = Category.objects.get_or_create(
            name=sub_category,
            main_category=main_category,
            category_type=1
        )
        return sub_category_obj
        # return int(sub_category_obj.id)


    def get_or_create_category(self, category, main_category, sub_main_category):
        from inventory.models import Category
        sub_main_category_created = self.get_or_create_sub_main_category(sub_main_category, main_category)
        category,is_created = Category.objects.get_or_create(
            name=category,
            main_category=main_category,
            sub_main_category=sub_main_category_created,
            category_type=2,
        )
        return [category.id]

    def get_or_create_tags(self, tags):
        splitted_tags = tags.split(",")
        tags = []
        for tag in splitted_tags:
            tag = Tags.objects.get_or_create(name=tag)
            tags.append(tag)
        return tags

    def get_variants_json(self, attribute_values, variants):
        data = {}
        data["mainState"] = json.loads(attribute_values)
        data["AttributesInputValue"] = json.loads(variants)
        return data

    def get_or_create_attribute(self, attributes):
        attributesList = attributes.split(",")
        for attribute in attributesList:
            try:
                attribute = Attributes.objects.get_or_create(name=attribute)
            except Exception as e:
                print_log(str(e))


    def create_product(self, item, image):
        data = {}
        main_category = item["main_category"]
        sub_main_category = item["sub_main_category"]
        is_new = item["is_new"]
        tags = item["tags"]
        item['thumb']=image.id
        item['sku']=str(item['sku'])
        item["category"] = self.get_or_create_category(
            item["category"], main_category, sub_main_category
        )
        item['main_category'] = int(item['main_category'])
        item["is_new_arrival"] = True if int(is_new) == 1 else False
        item['variant'] = json.loads(item['variant'])
        item['tags'] = json.loads(item['tags'])
        item.pop("is_new")
        item.pop("sub_main_category")
        from inventory.api.inventory.serializers import ProductSerializer
        print("item1212", item)
        serializer = ProductSerializer(data=item)
        if serializer.is_valid():
            print("sk__12",serializer.validated_data)
            print("sk__12","saving data")
            serializer.save()
        else:
            from rest_framework import serializers
            from inventory.models import Products
            found_sku = Products.objects.filter(sku=item['sku'])
            print("found_sku_sk__12", found_sku)
            print("found_sku_sk__12", Products.objects.filter(sku=item['sku']).count())
            found_sku[0].delete()
            raise serializers.ValidationError(serializer.errors)
        return 1

    def create_objects(self, result):
        index = 0
        for item in self.result:
            if item["name"] == "name":  # skip heading
                continue
            image = self.create_document(self.images[index])
            product = self.create_product(item, image)
            index += 1

    def bulk_import(self, file, filepath, request, *args, **kwargs):
        self.user = request.user
        uploaded_file_path = self.upload_file(file, filepath)
        wb = load_workbook(uploaded_file_path)
        sheet = wb["Sheet1"]
        max_column = sheet.max_column
        max_row = sheet.max_row
        self.result = []

        for row in range(1, max_row + 1):
            current_data = {}
            new_path = ""
            for column in range(1, max_column + 1):
                value = sheet.cell(row, column).value
                if row == 1:
                    self.heading.append(value)

                if column == 2 and row != 0:
                    try:
                        # if row!=max_row
                        image_loader = SheetImageLoader(sheet)
                        image = image_loader.get("B" + str(row + 1))
                        dir_path = self.img_filepath
                        pathlib.Path(settings.MEDIA_ROOT + f"/{dir_path}").mkdir(
                            parents=True, exist_ok=True
                        )
                        new_path = dir_path + f"{uuid.uuid4()}.PNG"
                        if image:
                            image.save(
                                settings.MEDIA_ROOT + f"/{new_path}",
                                "PNG",
                                optimize=False,
                                quality=100,
                            )
                        self.images.append(new_path)

                    except Exception as e:
                        print("thumb image issue",str(e))
                        default_path = f"default\placeholder.jpg"
                        self.images.append(default_path)

                else:
                    try:
                        current_data[self.heading[column - 1]] = value
                        current_data["row"] = row
                        current_data["column"] = column
                    except:
                        traceback.print_exc()
            self.result.append(current_data)
        self.create_objects(self.result[0])
        return 1


class CustomerBulkImport(BulkImportUtils):
    images = []
    result = []
    heading = []

    def create_customer(self,result):
        from inventory.models import Customer
        from inventory.api.inventory.serializers import CustomerSerializer

        for item in result:
            try:
                mobile = "+880"+str(int(item['mobile']))
            except:
                mobile = "+880"+item['mobile']
            item['mobile'] = mobile
            serializer = CustomerSerializer(data=item)
            if serializer.is_valid():
                serializer.save()
            else:
                from rest_framework import serializers
                raise serializers.ValidationError(serializer.errors)
        return 1



    
    def bulk_import(self, file, filepath, request, *args, **kwargs):
        uploaded_file_path = self.upload_file(file, filepath)
        wb = load_workbook(uploaded_file_path)
        sheet = wb["Sheet1"]
        max_column = sheet.max_column
        max_row = sheet.max_row
        self.result = []

        for row in range(1, max_row+1):
            current_data = {}
            new_path = ''

            for column in range(1, max_column+1):
                value = sheet.cell(row, column).value
                # print(value,end="  ")
                if row == 1:
                   self. heading.append(value)

                else:
                    try:
                        current_data[self.heading[column-1]] = value
                    except:
                        traceback.print_exc()
            self.result.append(current_data)
        self.result.pop(0)  # removing header values
        index = 0
        self.create_customer(self.result)
        return 1